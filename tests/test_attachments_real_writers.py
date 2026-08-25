"""Cross-validation against real spreadsheet writers, not just our own factory.

The factory in xlsx_factory.py writes inline strings; Excel, openpyxl and xlsxwriter write shared
strings, styled dates and their own quirks. These tests pin the reader against genuine output from
the two most common writers. They skip cleanly when the writers are not installed, so the
zero-external-dependency run path is untouched.
"""
from __future__ import annotations

import datetime as dt
import io

import pytest

from core.attachments import analyze_attachment, open_workbook

openpyxl = pytest.importorskip("openpyxl")


def _openpyxl_book() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["Customer", "Qty", "Posted", "Active", "Notes"])
    ws.append(["C-1", 5, dt.date(2026, 3, 1), True, "repeated note"])
    ws.append(["C-2", 2.5, dt.datetime(2026, 3, 2, 14, 30), False, "repeated note"])
    hidden = wb.create_sheet("HiddenData")
    hidden.sheet_state = "hidden"
    hidden.append(["Col"])
    hidden.append(["val"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_shared_strings_styled_dates_and_bools_from_openpyxl():
    """The factory cannot produce this file shape; a real writer can, and Excel does."""
    with open_workbook(io.BytesIO(_openpyxl_book())) as book:
        rows = list(book.sheets[0].rows())
        assert book.sheets[1].hidden is True

    assert [c.value for c in rows[0]] == ["Customer", "Qty", "Posted", "Active", "Notes"]
    r1 = {c.column: c for c in rows[1]}
    assert (r1[1].value, r1[1].kind) == ("C-1", "text")          # shared-strings path
    assert r1[3].kind == "date" and r1[3].value == dt.date(2026, 3, 1)
    assert (r1[4].value, r1[4].kind) == (True, "bool")
    r2 = {c.column: c for c in rows[2]}
    assert r2[3].value == dt.datetime(2026, 3, 2, 14, 30)        # time fraction survives
    assert r1[5].value == r2[5].value == "repeated note"          # string reuse decodes twice


def test_openpyxl_file_flows_through_the_full_analysis():
    rep = analyze_attachment(io.BytesIO(_openpyxl_book()), filename="real.xlsx",
                             slots={"data_fields": "customer, quantity, posted"})
    assert rep.fitness.verdict == "ready"
    assert any(f.code == "hidden_sheet_with_data" for f in rep.findings)


def test_rich_text_runs_concatenate():
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
    except ImportError:
        pytest.skip("openpyxl without rich_text support")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Header"])
    ws["A2"] = CellRichText(TextBlock(InlineFont(b=True), "bold "), "plain")
    buf = io.BytesIO()
    wb.save(buf)
    with open_workbook(io.BytesIO(buf.getvalue())) as book:
        rows = list(book.sheets[0].rows())
    assert rows[1][0].value == "bold plain"


def test_a_formula_without_a_cached_value_is_skipped_not_crashed():
    """openpyxl writes formulas with no cached result; Excel-saved files always cache one."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B"])
    ws.append([1, "=A2*2"])
    buf = io.BytesIO()
    wb.save(buf)
    with open_workbook(io.BytesIO(buf.getvalue())) as book:
        rows = list(book.sheets[0].rows())
    values = {c.column: c.value for c in rows[1]}
    assert values.get(1) == 1
    assert 2 not in values


def test_xlsxwriter_output_parses_identically():
    xlsxwriter = pytest.importorskip("xlsxwriter")
    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf, {"in_memory": True})
    ws = wb.add_worksheet("Data")
    datefmt = wb.add_format({"num_format": "yyyy-mm-dd"})
    for col, name in enumerate(["Customer", "Qty", "Posted"]):
        ws.write(0, col, name)
    ws.write(1, 0, "C-1"); ws.write(1, 1, 5)
    ws.write_datetime(1, 2, dt.datetime(2026, 3, 1), datefmt)
    wb.close()
    rep = analyze_attachment(io.BytesIO(buf.getvalue()), filename="xw.xlsx",
                             slots={"data_fields": "customer, quantity, posted"})
    assert rep.fitness.verdict == "ready"
    posted = next(c for c in rep.sheets[0].columns if c.header == "Posted")
    assert posted.dominant_kind() == "date"
